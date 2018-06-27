# -*- coding: utf-8 -*-
#================================================================================#
#  ECOLOG - Sistema Gerenciador de Banco de Dados para Levantamentos Ecológicos  #
#        ECOLOG - Database Management System for Ecological Surveys              #
#      Copyright (c) 1990-2016 Mauro J. Cavalcanti. All rights reserved.         #
#                                                                                #
#   Este programa é software livre; você pode redistribuí-lo e/ou                #
#   modificá-lo sob os termos da Licença Pública Geral GNU, conforme             #
#   publicada pela Free Software Foundation; tanto a versão 2 da                 #
#   Licença como (a seu critério) qualquer versão mais nova.                     #
#                                                                                # 
#   Este programa é distribuído na expectativa de ser útil, mas SEM              #
#   QUALQUER GARANTIA; sem mesmo a garantia implícita de                         #
#   COMERCIALIZAÇÃO ou de ADEQUAÇÃO A QUALQUER PROPÓSITO EM                      #
#   PARTICULAR. Consulte a Licença Pública Geral GNU para obter mais             #
#   detalhes.                                                                    #
#                                                                                #
#   This program is free software: you can redistribute it and/or                #
#   modify it under the terms of the GNU General Public License as published     #
#   by the Free Software Foundation, either version 2 of the License, or         #
#   version 3 of the License, or (at your option) any later version.             #
#                                                                                #
#   This program is distributed in the hope that it will be useful,              #
#   but WITHOUT ANY WARRANTY; without even the implied warranty of               #
#   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See                     #
#   the GNU General Public License for more details.                             #
#                                                                                #
#  Dependências / Dependencies:                                                  #
#    Python 2.6+  (www.python.org)                                               #
#    PyQt 4.8+  (www.riverbankcomputing.com/software/pyqt)                       #
#    openpyxl 2.0+ (openpyxl.readthedocs.org/en/2.0)                             #
#    ezodf 0.2+ (pythonhosted.org/ezodf)                                         #
#    xlwt 0.7+ (www.python-excel.org)                                            #                                           
#================================================================================#

import os, csv, codecs, zipfile, StringIO, cStringIO
import xlwt
import openpyxl
from ezodf import newdoc, Table

from Useful import (alfa, iif, to_float, to_int, unicode_to_ascii)
from Webservices import searchWoRMS

class UTF8Recoder:
    """
    Iterator that reads an encoded stream and reencodes the input to UTF-8
    """
    def __init__(self, f, encoding):
        self.reader = codecs.getreader(encoding)(f)

    def __iter__(self):
        return self

    def next(self):
        return self.reader.next().encode("utf-8")

class UnicodeReader:
    """
    A CSV reader which will iterate over lines in the CSV file "f",
    which is encoded in the given encoding.
    """

    def __init__(self, f, dialect=csv.excel, encoding="utf-8", **kwds):
        f = UTF8Recoder(f, encoding)
        self.reader = csv.reader(f, dialect=dialect, **kwds)

    def next(self):
        row = self.reader.next()
        return [unicode(s, "utf-8") for s in row]

    def __iter__(self):
        return self

class UnicodeWriter:
    """
    A CSV writer which will write rows to CSV file "f",
    which is encoded in the given encoding.
    """

    def __init__(self, f, dialect=csv.excel, encoding="utf-8", **kwds):
        # Redirect output to a queue
        self.queue = cStringIO.StringIO()
        self.writer = csv.writer(self.queue, dialect=dialect, **kwds)
        self.stream = f
        self.encoder = codecs.getincrementalencoder(encoding)()

    def writerow(self, row):
        self.writer.writerow([s.encode("utf-8") for s in row])
        # Fetch UTF-8 output from the queue ...
        data = self.queue.getvalue()
        data = data.decode("utf-8")
        # ... and reencode it into the target encoding
        data = self.encoder.encode(data)
        # write to the target stream
        self.stream.write(data)
        # empty queue
        self.queue.truncate(0)

    def writerows(self, rows):
        for row in rows:
            self.writerow(row)

def fromGBIF(filename, fmt):
    infile = os.path.basename(str(filename))
    outfile = os.path.splitext(infile)[0] + "_imported"
    
    fields = {'gbifID':0,'abstract':1,'accessRights':2,'accrualMethod':3,'accrualPeriodicity':4,'accrualPolicy':5,
            'alternative':6,'audience':7,'available':8,'bibliographicCitation':9,'conformsTo':10,'contributor':11,
            'coverage':12,'created':13,'creator':14,'date':15,'dateAccepted':16,'dateCopyrighted':17,
            'dateSubmitted':18,'description':19,'educationLevel':20,'extent':21,'format':22,'hasFormat':23,	
            'hasPart':24,'hasVersion':25,'identifier':26,'instructionalMethod':27,'isFormatOf':28,'isPartOf':29,
            'isReferencedBy':30,'isReplacedBy':31,'isRequiredBy':32,'isVersionOf':33,'issued':34,'language':35,
            'license':36,'mediator':37,'medium':38,'modified':39,'provenance':40,'publisher':41,'references':42,
            'relation':43,'replaces':44,'requires':45,'rights':46,'rightsHolder':47,'source':48,'spatial':49,
            'subject':50,'tableOfContents':51,'temporal':52,'title':53,'type':54,'valid':55,'acceptedNameUsage':56,
            'acceptedNameUsageID':57,'associatedOccurrences':58,'associatedReferences':59,'associatedSequences':60,
            'associatedTaxa':61,'basisOfRecord':62,'bed':63,'behavior':64,'catalogNumber':65,'class':66,
            'collectionCode':67,'collectionID':68,'continent':69,'countryCode':70,'county':71,
            'dataGeneralizations':72,'datasetID':73,'datasetName':74,'dateIdentified':75,'day':76,
            'decimalLatitude':77,'decimalLongitude':78,'disposition':79,'dynamicProperties':80,
            'earliestAgeOrLowestStage':81,'earliestEonOrLowestEonothem':82,'earliestEpochOrLowestSeries':83,
            'earliestEraOrLowestErathem':84,'earliestPeriodOrLowestSystem':85,'endDayOfYear':86,
            'establishmentMeans':87,'eventDate':88,'eventID':89,'eventRemarks':90,'eventTime':91,'family':92,
            'fieldNotes':93,'fieldNumber':94,'footprintSRS':95,'footprintSpatialFit':96,'footprintWKT':97,
            'formation':98,'genus':99,'geologicalContextID':100,'georeferencedDate':101,'georeferenceProtocol':102,
            'georeferenceRemarks':103,'georeferenceSources':104,'georeferenceVerificationStatus':105,
            'georeferencedBy':106,'group':107,'habitat':108,'higherClassification':109,'higherGeography':110,
            'higherGeographyID':111,'highestBiostratigraphicZone':112,'identificationID':113,
            'identificationQualifier':114,'identificationReferences':115,'identificationRemarks':116,
            'identificationVerificationStatus':117,'identifiedBy':118,'individualCount':119,'individualID':120,
            'informationWithheld':121,'infraspecificEpithet':122,'institutionCode':123,'institutionID':124,
            'island':125,'islandGroup':126,'kingdom':127,'latestAgeOrHighestStage':128,
            'latestEonOrHighestEonothem':129,'latestEpochOrHighestSeries':130,'latestEraOrHighestErathem':131,
            'latestPeriodOrHighestSystem':132,'lifeStage':133,'lithostratigraphicTerms':134,'locality':135,
            'locationAccordingTo':136,'locationID':137,'locationRemarks':138,'lowestBiostratigraphicZone':139,
            'materialSampleID':140,'maximumDistanceAboveSurfaceInMeters':141,'member':142,
            'minimumDistanceAboveSurfaceInMeters':143,'month':144,'municipality':145,'nameAccordingTo':146,
            'nameAccordingToID':147,'namePublishedIn':148,'namePublishedInID':149,'namePublishedInYear':150,
            'nomenclaturalCode':151,'nomenclaturalStatus':152,'occurrenceDetails':153,'occurrenceID':154,
            'occurrenceRemarks':155,'occurrenceStatus':156,'order':157,'originalNameUsage':158,
            'originalNameUsageID':159,'otherCatalogNumbers':160,'ownerInstitutionCode':161,'parentNameUsage':162,
            'parentNameUsageID':163,'phylum':164,'pointRadiusSpatialFit':165,'preparations':166,
            'previousIdentifications':167,'recordNumber':168,'recordedBy':169,'reproductiveCondition':170,
            'samplingEffort':171,'samplingProtocol':172,'scientificName':173,'scientificNameID':174,'sex':175,
            'specificEpithet':176,'startDayOfYear':177,'stateProvince':178,'subgenus':179,'taxonConceptID':180,
            'taxonID':181,'taxonRank':182,'taxonRemarks':183,'taxonomicStatus':184,'typeStatus':185,
            'verbatimCoordinateSystem':186,'verbatimDepth':187,'verbatimElevation':188,'verbatimEventDate':189,
            'verbatimLocality':190,'verbatimSRS':191,'verbatimTaxonRank':192,'vernacularName':193,'waterBody':194,
            'year':195,'datasetKey':196,'publishingCountry':197,'lastInterpreted':198,'coordinateAccuracy':199,
            'elevation':200,'elevationAccuracy':201,'depth':202,'depthAccuracy':203,'distanceAboveSurface':204,
            'distanceAboveSurfaceAccuracy':205,'issue':206,'mediaType':207,'hasCoordinate':208,
            'hasGeospatialIssues':209,'taxonKey':210,'kingdomKey':211,'phylumKey':212,'classKey':213,'orderKey':214,
            'familyKey':215,'genusKey':216,'subgenusKey':217,'speciesKey':218,'species':219,'genericName':220,
            'typifiedName':221,'protocol':222,'lastParsed':223,'lastCrawled':224}

    headers = ["Sample","Individual","Family","Scientific name","Collector name","Collector number",
                "Date collected","Locality","Latitude","Longitude","Altitude/Depth (m)","Kingdom","Phylum",
                "Class","Order","Sex","LifeStage","ReproductiveCondition","Behavior","Habitat","Notes"]
    
    if fmt == "csv":
        outfile += ".csv"
        writer = csv.writer(open(outfile, "wb"), delimiter=',', quotechar='"', quoting=csv.QUOTE_NONNUMERIC)
        writer.writerow(headers)
    elif fmt == "xls":
        outfile += ".xls"
        book = xlwt.Workbook()
        sheet1 = book.add_sheet("Dados GBIF")
        row1 = sheet1.row(0)
        for i in range(len(headers)):
            row1.write(i, headers[i])
    elif fmt == "xlsx":
        outfile += ".xlsx"
        book = openpyxl.Workbook()
        sheet1 = book.active
        sheet1.title = "Dados GBIF"
        for i in range(1, len(headers) + 1):
            sheet1.cell(row = 1, column = i).value = headers[i - 1]
    elif fmt == "ods":
        outfile += ".ods"
        book = newdoc(doctype="ods", filename=outfile)
        book.sheets.append(Table("Dados GBIF"))
        sheet1 = book.sheets[0]
        sheet1.append_columns(len(headers))
        for i in range(len(headers)):
            sheet1[0, i].set_value(headers[i])
    
    if infile.endswith(".zip"):
        zfile = zipfile.ZipFile(infile)
        data = StringIO.StringIO(zfile.read("occurrence.txt"))
        datafile = csv.reader(data, delimiter = '\t')
    elif infile.endswith(".txt"):
        datafile = csv.reader(open(infile, "rb"), delimiter = '\t')
    datafile.next()
    
    recnum = 0
    for row in datafile:
        if len(row[fields['earliestEpochOrLowestSeries']]) > 0 \
            or len(row[fields['earliestEraOrLowestErathem']]) > 0 \
            or len(row[fields['earliestPeriodOrLowestSystem']]) > 0:
            continue

        try:
            DateCollected = row[fields['eventDate']]
            InstitutionCode = row[fields['institutionCode']]
            CatalogNumber = row[fields['catalogNumber']]
            CollectorName = unicode_to_ascii(row[fields['recordedBy']])
            Sex = row[fields['sex']]
            LifeStage = row[fields['lifeStage']]
            ReproductiveCondition = row[fields['reproductiveCondition']]
            Behavior = row[fields['behavior']]
            Habitat = row[fields['habitat']]
            CollectorNumber = row[fields['type']]
            Notes = row[fields['eventRemarks']] + iif(len(row[fields['fieldNotes']]) > 0, ". " + row[fields['fieldNotes']], '')
            Locality = unicode_to_ascii(row[fields['locality']])
            Altitude = to_float(row[fields['elevation']])
            Depth = to_float(row[fields['depth']])
            Latitude = to_float(row[fields['decimalLatitude']])
            Longitude = to_float(row[fields['decimalLongitude']])
            ScientificName = unicode_to_ascii(row[fields['scientificName']])
            Kingdom = row[fields['kingdom']]
            Phylum = row[fields['phylum']]
            Class = row[fields['class']]
            Order = row[fields['order']]
            Family = row[fields['family']]
        except:
            continue
    
        if DateCollected.find('/') != -1:
            DateCollected = DateCollected.replace('/', '-')
    
        MyCollectionCode = InstitutionCode + CatalogNumber
        
        if Depth == 0.0:
            Elevation = Altitude
        elif Elevation == 0.0:
            Elevation = Depth
        
        if fmt == "csv":
            data = (recnum + 1, MyCollectionCode, Family, ScientificName, \
                CollectorName, CollectorNumber, DateCollected,\
                Locality, Latitude, Longitude, Elevation, \
                Kingdom, Phylum, Class, Order, \
                Sex, LifeStage, ReproductiveCondition, Behavior, Habitat, Notes)
            writer.writerow(data)
        elif fmt == "xls":
            sheet1.row(recnum + 1).write(0, recnum + 1)
            sheet1.row(recnum + 1).write(1, MyCollectionCode)
            sheet1.row(recnum + 1).write(2, Family)
            sheet1.row(recnum + 1).write(3, ScientificName)
            sheet1.row(recnum + 1).write(4, CollectorName)
            sheet1.row(recnum + 1).write(5, CollectorNumber)
            sheet1.row(recnum + 1).write(6, DateCollected)
            sheet1.row(recnum + 1).write(7, Locality)
            sheet1.row(recnum + 1).write(8, Latitude)
            sheet1.row(recnum + 1).write(9, Longitude)
            sheet1.row(recnum + 1).write(10, Elevation)
            sheet1.row(recnum + 1).write(11, Kingdom)
            sheet1.row(recnum + 1).write(12, Phylum)
            sheet1.row(recnum + 1).write(13, Class)
            sheet1.row(recnum + 1).write(14, Order)
            sheet1.row(recnum + 1).write(15, Sex)
            sheet1.row(recnum + 1).write(16, LifeStage)
            sheet1.row(recnum + 1).write(17, ReproductiveCondition)
            sheet1.row(recnum + 1).write(18, Behavior)
            sheet1.row(recnum + 1).write(19, Habitat)
            sheet1.row(recnum + 1).write(20, Notes)
        elif fmt == "xlsx":
            sheet1.cell(row=recnum + 2, column=1).value = recnum + 1
            sheet1.cell(row=recnum + 2, column=2).value = MyCollectionCode
            sheet1.cell(row=recnum + 2, column=3).value = Family
            sheet1.cell(row=recnum + 2, column=4).value = ScientificName
            sheet1.cell(row=recnum + 2, column=5).value = CollectorName
            sheet1.cell(row=recnum + 2, column=6).value = CollectorNumber
            sheet1.cell(row=recnum + 2, column=7).value = DateCollected
            sheet1.cell(row=recnum + 2, column=8).value = Locality
            sheet1.cell(row=recnum + 2, column=9).value = Latitude
            sheet1.cell(row=recnum + 2, column=10).value = Longitude
            sheet1.cell(row=recnum + 2, column=11).value = Elevation
            sheet1.cell(row=recnum + 2, column=12).value = Kingdom
            sheet1.cell(row=recnum + 2, column=13).value = Phylum
            sheet1.cell(row=recnum + 2, column=14).value = Class
            sheet1.cell(row=recnum + 2, column=15).value = Order
            sheet1.cell(row=recnum + 2, column=16).value = Sex
            sheet1.cell(row=recnum + 2, column=17).value = LifeStage
            sheet1.cell(row=recnum + 2, column=18).value = ReproductiveCondition
            sheet1.cell(row=recnum + 2, column=19).value = Behavior
            sheet1.cell(row=recnum + 2, column=20).value = Habitat
            sheet1.cell(row=recnum + 2, column=21).value = Notes
        elif fmt == "ods":
            sheet1.append_rows(1)
            sheet1[recnum + 1, 0].set_value(recnum + 1)
            sheet1[recnum + 1, 1].set_value(MyCollectionCode)
            sheet1[recnum + 1, 2].set_value(Family)
            sheet1[recnum + 1, 3].set_value(ScientificName)
            sheet1[recnum + 1, 4].set_value(CollectorName)
            sheet1[recnum + 1, 5].set_value(CollectorNumber)
            sheet1[recnum + 1, 6].set_value(DateCollected)
            sheet1[recnum + 1, 7].set_value(Locality)
            sheet1[recnum + 1, 8].set_value(Latitude)
            sheet1[recnum + 1, 9].set_value(Longitude)
            sheet1[recnum + 1, 10].set_value(Elevation)
            sheet1[recnum + 1, 11].set_value(Kingdom)
            sheet1[recnum + 1, 12].set_value(Phylum)
            sheet1[recnum + 1, 13].set_value(Class)
            sheet1[recnum + 1, 14].set_value(Order)
            sheet1[recnum + 1, 15].set_value(Sex)
            sheet1[recnum + 1, 16].set_value(LifeStage)
            sheet1[recnum + 1, 17].set_value(ReproductiveCondition)
            sheet1[recnum + 1, 18].set_value(Behavior)
            sheet1[recnum + 1, 19].set_value(Habitat)
            sheet1[recnum + 1, 20].set_value(Notes)
        recnum += 1
    
    if fmt == "xls":
        book.save(outfile)
    elif fmt == "xlsx":
        book.save(filename=outfile)
    elif fmt == "ods":
        book.save()
        
    return recnum, outfile

def fromOBIS(filename, fmt):
    infile = os.path.basename(str(filename))
    outfile = os.path.splitext(infile)[0] + "_imported"
    
    fields = {'id':0,'valid_id':1,'sname':2,'sauthor':3,'tname':4,'tauthor':5,'resource_id':6,'resname':7,
                'datecollected':8,'latitude':9,'longitude':10,'lifestage':11,'basisofrecord':12,
                'datelastcached':13,'dateprecision':14,'datelastmodified':15,'depth':16,'depthprecision':17,
                'temperature':18,'salinity':19,'nitrate':20,'oxygen':21,'phosphate':22,'silicate':23}
    
    headers = ["Sample","Individual","Family","Scientific name","Collector name","Collector number",
                "Date collected","Locality","Latitude","Longitude","Depth (m)","Kingdom","Phylum",
                "Class","Order","Sex","LifeStage","ReproductiveCondition","Behavior","Habitat","Notes"]
    
    if fmt == "csv":
        outfile += ".csv"
        writer = csv.writer(open(outfile, "wb"), delimiter=',', quotechar='"', quoting=csv.QUOTE_NONNUMERIC)
        writer.writerow(headers)
    elif fmt == "xls":
        outfile += ".xls"
        book = xlwt.Workbook()
        sheet1 = book.add_sheet("Dados OBIS")
        row1 = sheet1.row(0)
        for i in range(len(headers)):
            row1.write(i, headers[i])
    elif fmt == "xlsx":
        outfile += ".xlsx"
        book = openpyxl.Workbook()
        sheet1 = book.active
        sheet1.title = "Dados OBIS"
        for i in range(1, len(headers) + 1):
            sheet1.cell(row = 1, column = i).value = headers[i - 1]
    elif fmt == "ods":
        outfile += ".ods"
        book = newdoc(doctype="ods", filename=outfile)
        book.sheets.append(Table("Dados OBIS"))
        sheet1 = book.sheets[0]
        sheet1.append_columns(len(headers))
        for i in range(len(headers)):
            sheet1[0, i].set_value(headers[i])
            
    datafile = csv.reader(open(infile, "rb"), delimiter = ',')
    datafile.next()
    
    recnum = 0
    for row in datafile:
        Individual = row[fields['id']]
        ScientificName = row[fields['sname']]
        CollectorName = ""
        CollectorNumber = 0
        DateCollected = row[fields['datecollected']]
        Locality = ""
        Latitude = to_float(row[fields['latitude']])
        Longitude = to_float(row[fields['longitude']])
        LifeStage = row[fields['lifestage']]
        Depth = to_float(row[fields['depth']])
    
        if recnum == 0:
            try:
                record = searchWoRMS(ScientificName)
                Family = record["family"]
            except:
                Family = ""
                pass
                
        if fmt == "csv":
            data = (recnum + 1, Individual, Family, ScientificName, \
                CollectorName, CollectorNumber, DateCollected,\
                Locality, Latitude, Longitude, Depth, \
                LifeStage)
            writer.writerow(data)
        elif fmt == "xls":
            sheet1.row(recnum + 1).write(0, recnum + 1)
            sheet1.row(recnum + 1).write(1, Individual)
            sheet1.row(recnum + 1).write(2, Family)
            sheet1.row(recnum + 1).write(3, ScientificName)
            sheet1.row(recnum + 1).write(4, CollectorName)
            sheet1.row(recnum + 1).write(5, CollectorNumber)
            sheet1.row(recnum + 1).write(6, DateCollected)
            sheet1.row(recnum + 1).write(7, Locality)
            sheet1.row(recnum + 1).write(8, Latitude)
            sheet1.row(recnum + 1).write(9, Longitude)
            sheet1.row(recnum + 1).write(10, Depth)
            sheet1.row(recnum + 1).write(11, LifeStage)
        elif fmt == "xlsx":
            sheet1.cell(row=recnum + 2, column=1).value = recnum + 1
            sheet1.cell(row=recnum + 2, column=2).value = Individual
            sheet1.cell(row=recnum + 2, column=3).value = Family
            sheet1.cell(row=recnum + 2, column=4).value = ScientificName
            sheet1.cell(row=recnum + 2, column=5).value = CollectorName
            sheet1.cell(row=recnum + 2, column=6).value = CollectorNumber
            sheet1.cell(row=recnum + 2, column=7).value = DateCollected
            sheet1.cell(row=recnum + 2, column=8).value = Locality
            sheet1.cell(row=recnum + 2, column=9).value = Latitude
            sheet1.cell(row=recnum + 2, column=10).value = Longitude
            sheet1.cell(row=recnum + 2, column=11).value = Depth
            sheet1.cell(row=recnum + 2, column=12).value = LifeStage
        elif fmt == "ods":
            sheet1.append_rows(1)
            sheet1[recnum + 1, 0].set_value(recnum + 1)
            sheet1[recnum + 1, 1].set_value(Individual)
            sheet1[recnum + 1, 2].set_value(Family)
            sheet1[recnum + 1, 3].set_value(ScientificName)
            sheet1[recnum + 1, 4].set_value(CollectorName)
            sheet1[recnum + 1, 5].set_value(CollectorNumber)
            sheet1[recnum + 1, 6].set_value(DateCollected)
            sheet1[recnum + 1, 7].set_value(Locality)
            sheet1[recnum + 1, 8].set_value(Latitude)
            sheet1[recnum + 1, 9].set_value(Longitude)
            sheet1[recnum + 1, 10].set_value(Depth)
            sheet1[recnum + 1, 11].set_value(LifeStage)
        recnum += 1

    if fmt == "xls":
        book.save(outfile)
    elif fmt == "xlsx":
        book.save(filename=outfile)
    elif fmt == "ods":
        book.save()
        
    return recnum, outfile

def fromTEAM(filename, fmt, protocol):
    infile = os.path.basename(str(filename))
    outfile = os.path.splitext(infile)[0] + "_imported"
    
    if protocol == "Tree":
        fields = {"Id":0,"Observation Date":1,"Family":2,"Genus":3,"Species":4,
                "Names of Collectors":5,"Diameter":6,"POM Height":7,"New Diameter":8,
                "New POM Height":9,"Condition Codes":10,"Sampling Period":11,"Comments":12,
                "Data Level":13,"Sampling Unit Name":14,"Latitude":15,"Longitude":16,
                "Spatial Method":17,"Subplot Number":18,"1ha Plot X Coordinate":19,
                "1ha Plot Y Coordinate":20,"Tree Number":21,"Site Name":22,"1ha Plot Number":23,
                "Elevation":24,"Protocol Version":25,"Data Set Creator Institution":26,
                "Data Set Creator Scientist":27,"Data Set Contact":28}
        
        headers = ["Sample","Individual","Family","Scientific name","Names of Collectors",
            "Collector number","Observation Date","Site Name","Latitude","Longitude","Elevation (m)","Comments",
            "Diameter (cm)","POM Height (m)","Condition Codes"]
        
    elif protocol == "Liana":
        fields = {"Id":0,"Observation Date":1,"Family":2,"Genus":3,"Species":4,"Names of Collectors":5,
                "Diameter at 1.3m":6,"Max Diameter":7,"Max Diameter POM":8,"Condition Codes":9,
                "Location Codes":10,"Sampling Period":11,"Comments":12,"Data Level":13,"Sampling Unit Name":14,
                "Latitude":15,"Longitude":16,"Method":17,"Subplot Number":18,"1ha Plot X Coordinate":19,
                "1ha Plot Y Coordinate":20,"Tree Number":21,"Site Name":22,"1ha Plot Number":23,"Elevation":24,
                "Protocol Version":25,"Data Set Creator Institution":26,"Data Set Creator Scientist":27,
                "Data Set Contact":28}
        
        headers = ["Sample","Individual","Family","Scientific name","Names of Collectors",
            "Collector number","Observation Date","Site Name","Latitude","Longitude","Elevation (m)","Comments", 
            "Diameter at 1.3m (cm)","Max Diameter (cm)","Max Diameter POM (cm)","Condition Codes"]

    elif protocol == "Butterfly":
        fields = {"Id":0,"Date Captured":1,"Time Captured":2,"Family":3,"Subfamily":4,"Genus":5,
                "Species":6,"Subspecies":7,"Gender":8,"Date Identified":9,"Identifier Name":10,
                "Comments":11,"Sampling Period":12,"Envelope":13,"Recapture":14,"Data Level":15,
                "Sampling Unit Name":16,"Latitude":17,"Longitude":18,"Method":19,"Ima X Coordinate":20,
                "Ima Y Coordinate":21,"Stratum":22,"Trap Number":23,"Site Name":24,"1ha Plot Number":25,
                "Elevation":26,"Protocol Version":27,"Data Set Creator Institution":28,
                "Data Set Creator Scientist":29,"Data Set Contact":30}
        
        headers = ["Sample","Individual","Family","Scientific name","Names of Collectors",
            "Collector number","Date/Time Captured","Site Name","Latitude","Longitude","Elevation (m)","Comments",
            "Subfamily", "Gender", "Stratum"]
        
    elif protocol == "Avian":
        fields = {"Id":0,"Observation Date":1,"Observation Start Time":2,"Observation End Time":3, 
                "Family":4,"Genus":5,"Species":6,"Auditory":7,"Visual":8,"Distance":9,
                "Analog Sound Recording ID":10,"Digital Sound Recording ID":11,"Names of Observers":12,
                "Collection Code":13,"Comments":14,"Sampling Period":15,"Data Level":16,"Sampling Unit Name":17,
                "Latitude":18,"Longitude":19,"Method":20,"Block X Coordinate":21,"Block Y Coordinate":22,
                "Site Name":23,"1ha Plot Number":24,"Elevation":25,"Protocol Version":26,
                "Data Set Creator Institution":27,"Data Set Creator Scientist":28,
                "Data Set Contact":29}
        
        headers = ["Sample","Individual","Family","Scientific name","Names of Observers",
            "Collection Code","Observation Date/Time","Site Name","Latitude","Longitude","Elevation (m)","Comments",
            "Auditory", "Visual", "Distance Class"]
        
    elif protocol == "Primate":
        fields = {"Id":0,"Observation Date":1,"Observation Time":2,"Sampling Event":3,"Observers Name":4,
            "Family":5,"Genus":6,"Species":7,"Subspecies":8,"Number of Individuals":9,"Adult Males":10,
            "Adult Females":11,"Male Juveniles":12,"Female Juveniles":13,"Infants":14,"Nonidentifiable":15,
            "Distance to Primate":16,"Distance to Angle":17,"Comments":18,"Data Level":19,"Block X Coordinate":20,
            "Block Y Coordinate":21,"Sampling Unit Name":22,"Latitude":23,"Longitude":24,"Site Name":25,
            "1ha Plot Number":26,"Elevation":27,"Protocol Version":28,"Data Set Creator Institution":29,
            "Data Set Creator Scientist":30,"Data Set Contact":31}
                
        headers = ["Sample","Individual","Family","Scientific name","Names of Observers",
            "Collection number","Observation Date/Time","Site Name","Latitude","Longitude","Elevation (m)",
            "Comments","Number of Individuals","Adult Males","Adult Females","Male Juveniles",
            "Female Juveniles","Infants","Nonidentifiable","Distance to Primate (m)","Distance Angle (o)"]
        
    elif protocol == "Climate":
        fields = {"Id":0,"Observation Date":1,"Observation Time":2,"Direct Solar Radiation":3,
            "Total Solar Radiation":4,"Diffuse Solar Radiation":5,"Wind Speed":6,"Wind Direction":7,
            "Precipitation":8,"Dry Temperature":9,"Wet Temperature":10,"Relative Humidity":11,
            "Soil Moisture":12,"Photosynthetic Photon Flux Density":13,"Observer":14,"Comments":15,
            "Data Level":16,"Sampling Unit Name":17,"Latitude":18,"Longitud":19,"Protocol Version":20,
            "Site Name":21,"Data Set Creator Institution":22,"Data Set Creator Scientist":23,
            "Data Set Contact":24}
        
        headers = ["Sampling","Observer","Observation Date/Time","Site Name","Latitude","Longitude",
            "Comments","Direct Solar Radiation (W/m2)","Total Solar Radiation (W/m2)","Diffuse Solar Radiation (W/m2)",
            "Wind Speed (m/s)","Wind Direction (o)","Precipitation (mm)","Dry Temperature (C)","Wet Temperature (C)",
            "Relative Humidity (%)", "Soil Moisture (%)","Photosynthetic Photon Flux Density (micromol/m2.s)"]
    
    if fmt == "csv":
        outfile += ".csv"
        writer = UnicodeWriter(open(outfile, "wb"), delimiter=',', quotechar='"', quoting=csv.QUOTE_NONNUMERIC)
        writer.writerow(headers)
    elif fmt == "xls":
        outfile += ".xls"
        book = xlwt.Workbook()
        sheet1 = book.add_sheet("Dados TEAM")
        row1 = sheet1.row(0)
        for i in range(len(headers)):
            row1.write(i, headers[i])
    elif fmt == "xlsx":
        outfile += ".xlsx"
        book = openpyxl.Workbook()
        sheet1 = book.active
        sheet1.title = "Dados TEAM"
        for i in range(1, len(headers) + 1):
            sheet1.cell(row = 1, column = i).value = headers[i - 1]
    elif fmt == "ods":
        outfile += ".ods"
        book = newdoc(doctype="ods", filename=outfile)
        book.sheets.append(Table("Dados TEAM"))
        sheet1 = book.sheets[0]
        sheet1.append_columns(len(headers))
        for i in range(len(headers)):
            sheet1[0, i].set_value(headers[i])
            
    datafile = UnicodeReader(open(infile, "rb"), delimiter = ',', encoding="latin-1")
    datafile.next()
    
    recnum = 0
    for row in datafile:
        if protocol == "Tree":
            try:
                Individual = row[fields['Id']]
                DateCollected = row[fields['Observation Date']]
                Family = row[fields['Family']]
                ScientificName = row[fields['Genus']] + ' ' + row[fields['Species']]
                CollectorName = unicode_to_ascii(row[fields['Names of Collectors']])
                CollectorNumber = '0'
                Diameter = row[fields['Diameter']] 
                Height = row[fields['POM Height']] 
                ConditionCodes = row[fields['Condition Codes']]	
                Comments = row[fields['Comments']]
                Sample = row[fields['Sampling Unit Name']]
                Latitude = row[fields['Latitude']] 
                Longitude = row[fields['Longitude']] 
                Locality = unicode_to_ascii(row[fields['Site Name']])
                Elevation = row[fields['Elevation']] 
            except:
                continue
    
            if fmt == "csv":
                data = (Sample, Individual, Family, ScientificName, \
                        CollectorName, CollectorNumber, DateCollected,\
                        Locality, Latitude, Longitude, Elevation, Comments, \
                        Diameter, Height, ConditionCodes)
                writer.writerow(data)
            elif fmt == "xls":
                sheet1.row(recnum + 1).write(0, Sample)
                sheet1.row(recnum + 1).write(1, Individual)
                sheet1.row(recnum + 1).write(2, Family)
                sheet1.row(recnum + 1).write(3, ScientificName)
                sheet1.row(recnum + 1).write(4, CollectorName)
                sheet1.row(recnum + 1).write(5, CollectorNumber)
                sheet1.row(recnum + 1).write(6, DateCollected)
                sheet1.row(recnum + 1).write(7, Locality)
                sheet1.row(recnum + 1).write(8, to_float(Latitude))
                sheet1.row(recnum + 1).write(9, to_float(Longitude))
                sheet1.row(recnum + 1).write(10, to_float(Elevation))
                sheet1.row(recnum + 1).write(11, Comments)
                sheet1.row(recnum + 1).write(12, to_float(Diameter))
                sheet1.row(recnum + 1).write(13, to_float(Height))
                sheet1.row(recnum + 1).write(14, ConditionCodes)
            elif fmt == "xlsx":
                sheet1.cell(row=recnum + 2, column=1).value = recnum + 1
                sheet1.cell(row=recnum + 2, column=2).value = Individual
                sheet1.cell(row=recnum + 2, column=3).value = Family
                sheet1.cell(row=recnum + 2, column=4).value = ScientificName
                sheet1.cell(row=recnum + 2, column=5).value = CollectorName
                sheet1.cell(row=recnum + 2, column=6).value = CollectorNumber
                sheet1.cell(row=recnum + 2, column=7).value = DateCollected
                sheet1.cell(row=recnum + 2, column=8).value = Locality
                sheet1.cell(row=recnum + 2, column=9).value = to_float(Latitude)
                sheet1.cell(row=recnum + 2, column=10).value = to_float(Longitude)
                sheet1.cell(row=recnum + 2, column=11).value = to_float(Elevation)
                sheet1.cell(row=recnum + 2, column=12).value = Comments
                sheet1.cell(row=recnum + 2, column=13).value = to_float(Diameter)
                sheet1.cell(row=recnum + 2, column=14).value = to_float(Height)
                sheet1.cell(row=recnum + 2, column=15).value = ConditionCodes
            elif fmt == "ods":
                sheet1.append_rows(1)
                sheet1[recnum + 1, 0].set_value(recnum + 1)
                sheet1[recnum + 1, 1].set_value(Individual)
                sheet1[recnum + 1, 2].set_value(Family)
                sheet1[recnum + 1, 3].set_value(ScientificName)
                sheet1[recnum + 1, 4].set_value(CollectorName)
                sheet1[recnum + 1, 5].set_value(CollectorNumber)
                sheet1[recnum + 1, 6].set_value(DateCollected)
                sheet1[recnum + 1, 7].set_value(Locality)
                sheet1[recnum + 1, 8].set_value(to_float(Latitude))
                sheet1[recnum + 1, 9].set_value(to_float(Longitude))
                sheet1[recnum + 1, 10].set_value(to_float(Elevation))
                sheet1[recnum + 1, 11].set_value(Comments)
                sheet1[recnum + 1, 12].set_value(to_float(Diameter))
                sheet1[recnum + 1, 13].set_value(to_float(Height))
                sheet1[recnum + 1, 14].set_value(ConditionCodes)
        
        elif protocol == "Liana":
            try:
                Individual = row[fields['Id']]
                DateCollected = row[fields['Observation Date']]
                Family = row[fields['Family']]
                ScientificName = row[fields['Genus']] + ' ' + row[fields['Species']]
                CollectorName = unicode_to_ascii(row[fields['Names of Collectors']])
                CollectorNumber = '0'
                Diameter = row[fields['Diameter at 1.3m']] 
                MaxDiameter	= row[fields['Max Diameter']]
                MaxDiameterPOM = row[fields['Max Diameter POM']]
                ConditionCodes = row[fields['Condition Codes']]	
                Comments = row[fields['Comments']]
                Sample = row[fields['Sampling Unit Name']]
                Latitude = row[fields['Latitude']] 
                Longitude = row[fields['Longitude']]
                Locality = unicode_to_ascii(row[fields['Site Name']])
                Elevation = row[fields['Elevation']]
            except:
                continue
    
            if fmt == "csv":
                data = (Sample, Individual, Family, ScientificName, \
                        CollectorName, CollectorNumber, DateCollected,\
                        Locality, Latitude, Longitude, Elevation, Comments, \
                        Diameter, MaxDiameter, MaxDiameterPOM, ConditionCodes)
                writer.writerow(data)
            elif fmt == "xls":
                sheet1.row(recnum + 1).write(0, Sample)
                sheet1.row(recnum + 1).write(1, Individual)
                sheet1.row(recnum + 1).write(2, Family)
                sheet1.row(recnum + 1).write(3, ScientificName)
                sheet1.row(recnum + 1).write(4, CollectorName)
                sheet1.row(recnum + 1).write(5, CollectorNumber)
                sheet1.row(recnum + 1).write(6, DateCollected)
                sheet1.row(recnum + 1).write(7, Locality)
                sheet1.row(recnum + 1).write(8, to_float(Latitude))
                sheet1.row(recnum + 1).write(9, to_float(Longitude))
                sheet1.row(recnum + 1).write(10, to_float(Elevation))
                sheet1.row(recnum + 1).write(11, Comments)
                sheet1.row(recnum + 1).write(12, to_float(Diameter))
                sheet1.row(recnum + 1).write(13, to_float(MaxDiameter))
                sheet1.row(recnum + 1).write(14, to_float(MaxDiameterPOM))
                sheet1.row(recnum + 1).write(15, ConditionCodes)
            elif fmt == "xlsx":
                sheet1.cell(row=recnum + 2, column=1).value = recnum + 1
                sheet1.cell(row=recnum + 2, column=2).value = Individual
                sheet1.cell(row=recnum + 2, column=3).value = Family
                sheet1.cell(row=recnum + 2, column=4).value = ScientificName
                sheet1.cell(row=recnum + 2, column=5).value = CollectorName
                sheet1.cell(row=recnum + 2, column=6).value = CollectorNumber
                sheet1.cell(row=recnum + 2, column=7).value = DateCollected
                sheet1.cell(row=recnum + 2, column=8).value = Locality
                sheet1.cell(row=recnum + 2, column=9).value = to_float(Latitude)
                sheet1.cell(row=recnum + 2, column=10).value = to_float(Longitude)
                sheet1.cell(row=recnum + 2, column=11).value = to_float(Elevation)
                sheet1.cell(row=recnum + 2, column=12).value = Comments
                sheet1.cell(row=recnum + 2, column=13).value = to_float(Diameter)
                sheet1.cell(row=recnum + 2, column=14).value = to_float(MaxDiameter)
                sheet1.cell(row=recnum + 2, column=15).value = to_float(MaxDiameterPOM)
                sheet1.cell(row=recnum + 2, column=16).value = ConditionCodes
            elif fmt == "ods":
                sheet1.append_rows(1)
                sheet1[recnum + 1, 0].set_value(recnum + 1)
                sheet1[recnum + 1, 1].set_value(Individual)
                sheet1[recnum + 1, 2].set_value(Family)
                sheet1[recnum + 1, 3].set_value(ScientificName)
                sheet1[recnum + 1, 4].set_value(CollectorName)
                sheet1[recnum + 1, 5].set_value(CollectorNumber)
                sheet1[recnum + 1, 6].set_value(DateCollected)
                sheet1[recnum + 1, 7].set_value(Locality)
                sheet1[recnum + 1, 8].set_value(to_float(Latitude))
                sheet1[recnum + 1, 9].set_value(to_float(Longitude))
                sheet1[recnum + 1, 10].set_value(to_float(Elevation))
                sheet1[recnum + 1, 11].set_value(Comments)
                sheet1[recnum + 1, 12].set_value(to_float(Diameter))
                sheet1[recnum + 1, 13].set_value(to_float(MaxDiameter))
                sheet1[recnum + 1, 14].set_value(to_float(MaxDiameterPOM))
                sheet1[recnum + 1, 15].set_value(ConditionCodes)
    
        elif protocol == "Butterfly":
            try:
                Individual = row[fields['Id']]
                DateCollected = row[fields['Date Captured']] + ' ' + row[fields['Time Captured']]
                Family = row[fields['Family']]
                Subfamily = row[fields['Subfamily']]
                ScientificName = row[fields['Genus']] + ' ' + row[fields['Species']] + \
                                iif(len(row[fields['Subspecies']]) > 0, ' ' + row[fields['Subspecies']], '')
                Gender = row[fields['Gender']]
                CollectorName = ""
                CollectorNumber = '0'
                Comments = row[fields['Comments']]
                Sample = row[fields['Sampling Unit Name']]
                Latitude = row[fields['Latitude']]
                Longitude = row[fields['Longitude']]
                Stratum = row[fields['Stratum']]
                Locality = unicode_to_ascii(row[fields['Site Name']])
                Elevation = row[fields['Elevation']]
            except:
                continue
    
            if fmt == "csv":
                data = (Sample, Individual, Family, ScientificName, \
                        CollectorName, CollectorNumber, DateCollected,\
                        Locality, Latitude, Longitude, Elevation, Comments, \
                        Subfamily, Gender, Stratum)
                writer.writerow(data)
            elif fmt == "xls":
                sheet1.row(recnum + 1).write(0, Sample)
                sheet1.row(recnum + 1).write(1, Individual)
                sheet1.row(recnum + 1).write(2, Family)
                sheet1.row(recnum + 1).write(3, ScientificName)
                sheet1.row(recnum + 1).write(4, CollectorName)
                sheet1.row(recnum + 1).write(5, CollectorNumber)
                sheet1.row(recnum + 1).write(6, DateCollected)
                sheet1.row(recnum + 1).write(7, Locality)
                sheet1.row(recnum + 1).write(8, to_float(Latitude))
                sheet1.row(recnum + 1).write(9, to_float(Longitude))
                sheet1.row(recnum + 1).write(10, to_float(Elevation))
                sheet1.row(recnum + 1).write(11, Comments)
                sheet1.row(recnum + 1).write(12, Subfamily)
                sheet1.row(recnum + 1).write(13, Gender)
                sheet1.row(recnum + 1).write(14, Stratum)
            elif fmt == "xlsx":
                sheet1.cell(row=recnum + 2, column=1).value = recnum + 1
                sheet1.cell(row=recnum + 2, column=2).value = Individual
                sheet1.cell(row=recnum + 2, column=3).value = Family
                sheet1.cell(row=recnum + 2, column=4).value = ScientificName
                sheet1.cell(row=recnum + 2, column=5).value = CollectorName
                sheet1.cell(row=recnum + 2, column=6).value = CollectorNumber
                sheet1.cell(row=recnum + 2, column=7).value = DateCollected
                sheet1.cell(row=recnum + 2, column=8).value = Locality
                sheet1.cell(row=recnum + 2, column=9).value = to_float(Latitude)
                sheet1.cell(row=recnum + 2, column=10).value = to_float(Longitude)
                sheet1.cell(row=recnum + 2, column=11).value = to_float(Elevation)
                sheet1.cell(row=recnum + 2, column=12).value = Comments
                sheet1.cell(row=recnum + 2, column=13).value = Subfamily
                sheet1.cell(row=recnum + 2, column=14).value = Gender
                sheet1.cell(row=recnum + 2, column=15).value = Stratum
            elif fmt == "ods":
                sheet1.append_rows(1)
                sheet1[recnum + 1, 0].set_value(recnum + 1)
                sheet1[recnum + 1, 1].set_value(Individual)
                sheet1[recnum + 1, 2].set_value(Family)
                sheet1[recnum + 1, 3].set_value(ScientificName)
                sheet1[recnum + 1, 4].set_value(CollectorName)
                sheet1[recnum + 1, 5].set_value(CollectorNumber)
                sheet1[recnum + 1, 6].set_value(DateCollected)
                sheet1[recnum + 1, 7].set_value(Locality)
                sheet1[recnum + 1, 8].set_value(to_float(Latitude))
                sheet1[recnum + 1, 9].set_value(to_float(Longitude))
                sheet1[recnum + 1, 10].set_value(to_float(Elevation))
                sheet1[recnum + 1, 11].set_value(Comments)
                sheet1[recnum + 1, 12].set_value(Subfamily)
                sheet1[recnum + 1, 13].set_value(Gender)
                sheet1[recnum + 1, 14].set_value(Stratum)

        elif protocol == "Avian":
            try:
                Individual = row[fields['Id']]
                DateCollected = row[fields['Observation Date']] + ' ' + row[fields['Observation Start Time']]
                Family = row[fields['Family']]
                ScientificName = row[fields['Genus']] + ' ' + row[fields['Species']]
                Auditory = row[fields['Auditory']]
                Visual = row[fields['Visual']]
                Distance = row[fields['Distance']]
                CollectorName = unicode_to_ascii(row[fields['Names of Observers']])
                CollectorNumber = row[fields['Collection Code']]
                Comments = row[fields['Comments']]
                Sample = row[fields['Sampling Unit Name']]
                Latitude = row[fields['Latitude']]
                Longitude = row[fields['Longitude']]
                Locality = unicode_to_ascii(row[fields['Site Name']])
                Elevation = row[fields['Elevation']]
            except:
                continue
    
            if fmt == "csv":
                data = (Sample, Individual, Family, ScientificName, \
                        CollectorName, CollectorNumber, DateCollected,\
                        Locality, Latitude, Longitude, Elevation, Comments, \
                        Auditory, Visual, Distance)
                writer.writerow(data)
            elif fmt == "xls":
                sheet1.row(recnum + 1).write(0, Sample)
                sheet1.row(recnum + 1).write(1, Individual)
                sheet1.row(recnum + 1).write(2, Family)
                sheet1.row(recnum + 1).write(3, ScientificName)
                sheet1.row(recnum + 1).write(4, CollectorName)
                sheet1.row(recnum + 1).write(5, CollectorNumber)
                sheet1.row(recnum + 1).write(6, DateCollected)
                sheet1.row(recnum + 1).write(7, Locality)
                sheet1.row(recnum + 1).write(8, to_float(Latitude))
                sheet1.row(recnum + 1).write(9, to_float(Longitude))
                sheet1.row(recnum + 1).write(10, to_float(Elevation))
                sheet1.row(recnum + 1).write(11, Comments)
                sheet1.row(recnum + 1).write(12, Auditory)
                sheet1.row(recnum + 1).write(13, Visual)
                sheet1.row(recnum + 1).write(14, to_int(Distance))
            elif fmt == "xlsx":
                sheet1.cell(row=recnum + 2, column=1).value = recnum + 1
                sheet1.cell(row=recnum + 2, column=2).value = Individual
                sheet1.cell(row=recnum + 2, column=3).value = Family
                sheet1.cell(row=recnum + 2, column=4).value = ScientificName
                sheet1.cell(row=recnum + 2, column=5).value = CollectorName
                sheet1.cell(row=recnum + 2, column=6).value = CollectorNumber
                sheet1.cell(row=recnum + 2, column=7).value = DateCollected
                sheet1.cell(row=recnum + 2, column=8).value = Locality
                sheet1.cell(row=recnum + 2, column=9).value = to_float(Latitude)
                sheet1.cell(row=recnum + 2, column=10).value = to_float(Longitude)
                sheet1.cell(row=recnum + 2, column=11).value = to_float(Elevation)
                sheet1.cell(row=recnum + 2, column=12).value = Comments
                sheet1.cell(row=recnum + 2, column=13).value = Auditory
                sheet1.cell(row=recnum + 2, column=14).value = Visual
                sheet1.cell(row=recnum + 2, column=15).value = to_int(Distance)
            elif fmt == "ods":
                sheet1.append_rows(1)
                sheet1[recnum + 1, 0].set_value(recnum + 1)
                sheet1[recnum + 1, 1].set_value(Individual)
                sheet1[recnum + 1, 2].set_value(Family)
                sheet1[recnum + 1, 3].set_value(ScientificName)
                sheet1[recnum + 1, 4].set_value(CollectorName)
                sheet1[recnum + 1, 5].set_value(CollectorNumber)
                sheet1[recnum + 1, 6].set_value(DateCollected)
                sheet1[recnum + 1, 7].set_value(Locality)
                sheet1[recnum + 1, 8].set_value(to_float(Latitude))
                sheet1[recnum + 1, 9].set_value(to_float(Longitude))
                sheet1[recnum + 1, 10].set_value(to_float(Elevation))
                sheet1[recnum + 1, 11].set_value(Comments)
                sheet1[recnum + 1, 12].set_value(Auditory)
                sheet1[recnum + 1, 13].set_value(Visual)
                sheet1[recnum + 1, 14].set_value(to_int(Distance))
            
        elif protocol == "Primate":
            try:
                Individual = row[fields['Id']]
                DateCollected = row[fields['Observation Date']] + ' ' + row[fields['Observation Time']]
                CollectorName = unicode_to_ascii(row[fields['Observers Name']])
                CollectorNumber = '0'
                Family = row[fields['Family']]
                ScientificName = row[fields['Genus']] + ' ' + row[fields['Species']] + \
                                iif(len(row[fields['Subspecies']]) > 0, row[fields['Subspecies']], '')
                NumberOfIndividuals = row[fields['Number of Individuals']]
                AdultMales = row[fields['Adult Males']]
                AdultFemales = row[fields['Adult Females']]
                MaleJuveniles = row[fields['Male Juveniles']]
                FemaleJuveniles = row[fields['Female Juveniles']]
                Infants = row[fields['Infants']]
                Nonidentifiable = row[fields['Nonidentifiable']]
                DistanceToPrimate = row[fields['Distance to Primate']]
                DistanceToAngle = row[fields['Distance to Angle']]
                Comments = row[fields['Comments']]
                Sample = row[fields['Sampling Unit Name']]
                Latitude = row[fields['Latitude']] 
                Longitude = row[fields['Longitude']] 
                Locality = unicode_to_ascii(row[25])
                Elevation = row[fields['Elevation']]
            except:
                continue
            
            if fmt == "csv":
                data = (Sample, Individual, Family, ScientificName, \
                        CollectorName, CollectorNumber, DateCollected,\
                        Locality, Latitude, Longitude, Elevation, Comments, \
                        NumberOfIndividuals, AdultMales, AdultFemales, \
                        MaleJuveniles, FemaleJuveniles, Infants, Nonidentifiable, \
                        DistanceToPrimate, DistanceToAngle)
                writer.writerow(data)
            elif fmt == "xls":
                sheet1.row(recnum + 1).write(0, Sample)
                sheet1.row(recnum + 1).write(1, Individual)
                sheet1.row(recnum + 1).write(2, Family)
                sheet1.row(recnum + 1).write(3, ScientificName)
                sheet1.row(recnum + 1).write(4, CollectorName)
                sheet1.row(recnum + 1).write(5, CollectorNumber)
                sheet1.row(recnum + 1).write(6, DateCollected)
                sheet1.row(recnum + 1).write(7, Locality)
                sheet1.row(recnum + 1).write(8, to_float(Latitude))
                sheet1.row(recnum + 1).write(9, to_float(Longitude))
                sheet1.row(recnum + 1).write(10, to_float(Elevation))
                sheet1.row(recnum + 1).write(11, Comments)
                sheet1.row(recnum + 1).write(12, to_int(NumberOfIndividuals))
                sheet1.row(recnum + 1).write(13, to_int(AdultMales))
                sheet1.row(recnum + 1).write(14, to_int(AdultFemales))
                sheet1.row(recnum + 1).write(15, to_int(MaleJuveniles))
                sheet1.row(recnum + 1).write(16, to_int(FemaleJuveniles))
                sheet1.row(recnum + 1).write(17, to_int(Infants))
                sheet1.row(recnum + 1).write(18, to_int(Nonidentifiable))
                sheet1.row(recnum + 1).write(19, to_float(DistanceToPrimate))
                sheet1.row(recnum + 1).write(20, to_float(DistancetoAngle))
            elif fmt == "xlsx":
                sheet1.cell(row=recnum + 2, column=1).value = recnum + 1
                sheet1.cell(row=recnum + 2, column=2).value = Individual
                sheet1.cell(row=recnum + 2, column=3).value = Family
                sheet1.cell(row=recnum + 2, column=4).value = ScientificName
                sheet1.cell(row=recnum + 2, column=5).value = CollectorName
                sheet1.cell(row=recnum + 2, column=6).value = CollectorNumber
                sheet1.cell(row=recnum + 2, column=7).value = DateCollected
                sheet1.cell(row=recnum + 2, column=8).value = Locality
                sheet1.cell(row=recnum + 2, column=9).value = to_float(Latitude)
                sheet1.cell(row=recnum + 2, column=10).value = to_float(Longitude)
                sheet1.cell(row=recnum + 2, column=11).value = to_float(Elevation)
                sheet1.cell(row=recnum + 2, column=12).value = Comments
                sheet1.cell(row=recnum + 2, column=13).value = to_int(NumberOfIndividuals)
                sheet1.cell(row=recnum + 2, column=14).value = to_int(AdultMales)
                sheet1.cell(row=recnum + 2, column=15).value = to_int(AdultFemales)
                sheet1.cell(row=recnum + 2, column=16).value = to_int(MaleJuveniles)
                sheet1.cell(row=recnum + 2, column=17).value = to_int(FemaleJuveniles)
                sheet1.cell(row=recnum + 2, column=18).value = to_int(Infants)
                sheet1.cell(row=recnum + 2, column=19).value = to_int(Nonidentifiable)
                sheet1.cell(row=recnum + 2, column=20).value = to_float(DistanceToPrimate)
                sheet1.cell(row=recnum + 2, column=21).value = to_float(DistanceToAngle)
            elif fmt == "ods":
                sheet1.append_rows(1)
                sheet1[recnum + 1, 0].set_value(recnum + 1)
                sheet1[recnum + 1, 1].set_value(Individual)
                sheet1[recnum + 1, 2].set_value(Family)
                sheet1[recnum + 1, 3].set_value(ScientificName)
                sheet1[recnum + 1, 4].set_value(CollectorName)
                sheet1[recnum + 1, 5].set_value(CollectorNumber)
                sheet1[recnum + 1, 6].set_value(DateCollected)
                sheet1[recnum + 1, 7].set_value(Locality)
                sheet1[recnum + 1, 8].set_value(to_float(Latitude))
                sheet1[recnum + 1, 9].set_value(to_float(Longitude))
                sheet1[recnum + 1, 10].set_value(to_float(Elevation))
                sheet1[recnum + 1, 11].set_value(Comments)
                sheet1[recnum + 1, 12].set_value(to_int(NumberOfIndividuals))
                sheet1[recnum + 1, 13].set_value(to_int(AdultMales))
                sheet1[recnum + 1, 14].set_value(to_int(AdultFemales))
                sheet1[recnum + 1, 15].set_value(to_int(MaleJuveniles))
                sheet1[recnum + 1, 16].set_value(to_int(FemaleJuveniles))
                sheet1[recnum + 1, 17].set_value(to_int(Infants))
                sheet1[recnum + 1, 18].set_value(to_int(Nonidentifiable))
                sheet1[recnum + 1, 19].set_value(to_float(DistanceToPrimate))
                sheet1[recnum + 1, 20].set_value(to_float(DistanceToAngle))
            
        elif protocol == "Climate":
            try:
                ObservationDate = row[fields['Observation Date']] + ' ' + row[fields['Observation Time']]
                DirectSolarRadiation = row[fields['Direct Solar Radiation']]
                TotalSolarRadiation	= row[fields['Total Solar Radiation']]
                DiffuseSolarRadiation = row[fields['Diffuse Solar Radiation']]
                WindSpeed = row[fields['Wind Speed']]
                WindDirection = row[fields['Wind Direction']]
                Precipitation = row[fields['Precipitation']]	
                DryTemperature = row[fields['Dry Temperature']]
                WetTemperature = row[fields['Wet Temperature']]
                RelativeHumidity = row[fields['Relative Humidity']]
                SoilMoisture = row[fields['Soil Moisture']]
                PhotosyntheticPhotonFluxDensity	= row[fields['Photosynthetic Photon Flux Density']]
                Observer = unicode_to_ascii(row[fields['Observer']])
                Comments = row[fields['Comments']]
                Sample = row[fields['Sampling Unit Name']]
                Latitude = row[fields['Latitude']]
                Longitude = row[fields['Longitud']]
                Locality = unicode_to_ascii(row[fields['Site Name']])
            except:
                continue
    
            if fmt == "csv":
                data = (Sample, Observer, ObservationDate, Locality, \
                        Latitude, Longitude, Comments, \
                        DirectSolarRadiation, TotalSolarRadiation, DiffuseSolarRadiation, \
                        WindSpeed, WindDirection, Precipitation, DryTemperature, WetTemperature,	\
                        RelativeHumidity, SoilMoisture, PhotosyntheticPhotonFluxDensity)
                writer.writerow(data)
            elif fmt == "xls":
                sheet1.row(recnum + 1).write(0, Sample)
                sheet1.row(recnum + 1).write(1, Observer)
                sheet1.row(recnum + 1).write(2, ObservationDate)
                sheet1.row(recnum + 1).write(3, Locality)
                sheet1.row(recnum + 1).write(4, to_float(Latitude))
                sheet1.row(recnum + 1).write(5, to_float(Longitude))
                sheet1.row(recnum + 1).write(6, Comments)
                sheet1.row(recnum + 1).write(7, to_float(DirectSolarRadiation))
                sheet1.row(recnum + 1).write(8, to_float(TotalSolarRadiation))
                sheet1.row(recnum + 1).write(9, to_float(DiffuseSolarRadiation))
                sheet1.row(recnum + 1).write(10, to_float(WindSpeed))
                sheet1.row(recnum + 1).write(11, to_float(WindDirection))
                sheet1.row(recnum + 1).write(12, to_float(Precipitation))
                sheet1.row(recnum + 1).write(13, to_float(DryTemperature))
                sheet1.row(recnum + 1).write(14, to_float(WetTemperature))
                sheet1.row(recnum + 1).write(15, to_float(RelativeHumidity))
                sheet1.row(recnum + 1).write(16, to_float(SoilMoisture))
                sheet1.row(recnum + 1).write(17, to_float(PhotosyntheticPhotonFluxDensity))
            elif fmt == "xlsx":
                sheet1.cell(row=recnum + 2, column=1).value = Sample
                sheet1.cell(row=recnum + 2, column=2).value = Observer
                sheet1.cell(row=recnum + 2, column=3).value = ObservationDate
                sheet1.cell(row=recnum + 2, column=4).value = Locality
                sheet1.cell(row=recnum + 2, column=5).value = to_float(Latitude)
                sheet1.cell(row=recnum + 2, column=6).value = to_float(Longitude)
                sheet1.cell(row=recnum + 2, column=7).value = Comments
                sheet1.cell(row=recnum + 2, column=8).value = to_float(DirectSolarRadiation)
                sheet1.cell(row=recnum + 2, column=9).value = to_float(TotalSolarRadiation)
                sheet1.cell(row=recnum + 2, column=10).value = to_float(DiffuseSolarRadiation)
                sheet1.cell(row=recnum + 2, column=11).value = to_float(WindSpeed)
                sheet1.cell(row=recnum + 2, column=12).value = to_float(WindDirection)
                sheet1.cell(row=recnum + 2, column=13).value = to_float(Precipitation)
                sheet1.cell(row=recnum + 2, column=14).value = to_float(DryTemperature)
                sheet1.cell(row=recnum + 2, column=15).value = to_float(WetTemperature)
                sheet1.cell(row=recnum + 2, column=16).value = to_float(RelativeHumidity)
                sheet1.cell(row=recnum + 2, column=17).value = to_float(SoilMoisture)
                sheet1.cell(row=recnum + 2, column=18).value = to_float(PhotosyntheticPhotonFluxDensity)
            elif fmt == "ods":
                sheet1.append_rows(1)
                sheet1[recnum + 1, 0].set_value(Sample)
                sheet1[recnum + 1, 1].set_value(Observer)
                sheet1[recnum + 1, 2].set_value(ObservationDate)
                sheet1[recnum + 1, 3].set_value(Locality)
                sheet1[recnum + 1, 4].set_value(to_float(Latitude))
                sheet1[recnum + 1, 5].set_value(to_float(Longitude))
                sheet1[recnum + 1, 6].set_value(Comments)
                sheet1[recnum + 1, 7].set_value(to_float(DirectSolarRadiation))
                sheet1[recnum + 1, 8].set_value(to_float(TotalSolarRadiation))
                sheet1[recnum + 1, 9].set_value(to_float(DiffuseSolarRadiation))
                sheet1[recnum + 1, 10].set_value(to_float(WindSpeed))
                sheet1[recnum + 1, 11].set_value(to_float(WindDirection))
                sheet1[recnum + 1, 12].set_value(to_float(Precipitation))
                sheet1[recnum + 1, 13].set_value(to_float(DryTemperature))
                sheet1[recnum + 1, 14].set_value(to_float(WetTemperature))
                sheet1[recnum + 1, 15].set_value(to_float(RelativeHumidity))
                sheet1[recnum + 1, 16].set_value(to_float(SoilMoisture))
                sheet1[recnum + 1, 17].set_value(to_float(PhotosyntheticPhotonFluxDensity))
        recnum += 1
    
    if fmt == "xls":
        book.save(outfile)
    elif fmt == "xlsx":
        book.save(filename=outfile)
    elif fmt == "ods":
        book.save()
        
    return recnum, outfile

def fromVertNet(filename, fmt):
    infile = os.path.basename(str(filename))
    outfile = os.path.splitext(infile)[0] + "_imported"
    
    fields = {'datasource_and_rights':0,'type':1,'modified':2,'language':3,'rights':4,'rightsholder':5,
                'accessrights':6,'bibliographiccitation':7,'references':8,'institutionid':9,'collectionid':10,
                'datasetid':11,'institutioncode':12,'collectioncode':13,'datasetname':14,'ownerinstitutioncode':15,
                'basisofrecord':16,'informationwithheld':17,'datageneralizations':18,'dynamicproperties':19,
                'occurrenceid':20,'catalognumber':21,'occurrenceremarks':22,'recordnumber':23,'recordedby':24,
                'individualid':25,'individualcount':26,'sex':27,'lifestage':28,'reproductivecondition':29,
                'behavior':30,'establishmentmeans':31,'occurrencestatus':32,'preparations':33,'disposition':34,
                'othercatalognumbers':35,'previousidentifications':36,'associatedmedia':37,'associatedreferences':38,
                'associatedoccurrences':40,'associatedsequences':41,'associatedtaxa':42,'eventid':43,
                'samplingprotocol':44,'samplingeffort':45,'eventdate':47,'eventtime':48,'startdayofyear':49,
                'enddayofyear':50,'year':51,'month':52,'day':53,'verbatimeventdate':54,'habitat':55,'fieldnumber':56,
                'fieldnotes':57,'eventremarks':58,'locationid':59,'highergeographyid':60,'highergeography':61,
                'continent':62,'waterbody':63,'islandgroup':64,'island':65,'country':66,'countrycode':67,
                'stateprovince':68,'county':69,'municipality':70,'locality':71,'verbatimlocality':72,
                'verbatimelevation':73,'minimumelevationinmeters':74,'maximumelevationinmeters':75,
                'verbatimdepth':76,'minimumdepthinmeters':77,'maximumdepthinmeters':78,
                'minimumdistanceabovesurfaceinmeters':79,'maximumdistanceabovesurfaceinmeters':80,
                'locationaccordingto':81,'locationremarks':82,'verbatimcoordinates':83,'verbatimlatitude':84,
                'verbatimlongitude':85,'verbatimcoordinatesystem':86,'verbatimsrs':87,'decimallatitude':88,
                'decimallongitude':89,'geodeticdatum':90,'coordinateuncertaintyinmeters':91,'coordinateprecision':93,
                'pointradiusspatialfit':94,'footprintwkt':95,'footprintsrs':96,'footprintspatialfit':97,
                'georeferencedby':98,'georeferenceddate':99,'georeferenceprotocol':100,'georeferencesources':101,
                'georeferenceverificationstatus':102,'georeferenceremarks':103,'geologicalcontextid':104,
                'earliesteonorlowesteonothem':105,'latesteonorhighesteonothem':106,'earliesteraorlowesterathem':107,
                'latesteraorhighesterathem':108,'earliestperiodorlowestsystem':109,'latestperiodorhighestsystem':110,
                'earliestepochorlowestseries':111,'latestepochorhighestseries':112,'earliestageorloweststage':113,
                'latestageorhigheststage':114,'lowestbiostratigraphiczone':115,'highestbiostratigraphiczone':116,
                'lithostratigraphicterms':117,'group':118,'formation':119,'member':120,'bed':121,'identificationid':122,
                'identifiedby':123,'dateidentified':124,'identificationreferences':125,
                'identificationverificationstatus':126,'identificationremarks':127,'identificationqualifier':128,
                'typestatus':129,'taxonid':130,'scientificnameid':131,'acceptednameusageid':132,
                'parentnameusageid':133,'originalnameusageid':134,'nameaccordingtoid':135,'namepublishedinid':136,
                'taxonconceptid':137,'scientificname':138,'acceptednameusage':139,'parentnameusage':140,
                'originalnameusage':141,'nameaccordingto':142,'namepublishedin':143,'namepublishedinyear':144,
                'higherclassification':145,'kingdom':146,'phylum':147,'class':148,'order':149,'family':150,
                'genus':151,'subgenus':152,'specificepithet':153,'infraspecificepithet':154,'taxonrank':155,
                'verbatimtaxonrank':156,'scientificnameauthorship':157,'vernacularname':158,'nomenclaturalcode':159,
                'taxonomicstatus':160,'nomenclaturalstatus':161,'taxonremarks':162}
    
    headers = ["Sample","Individual","Family","Scientific name","Collector name","Collector number",
                "Date collected","Locality","Latitude","Longitude","Altitude/Depth (m)","Kingdom","Phylum",
                "Class","Order","Sex","LifeStage","ReproductiveCondition","Behavior","Habitat","Notes"]
    
    if fmt == "csv":
        outfile += ".csv"
        writer = csv.writer(open(outfile, "wb"), delimiter=',', quotechar='"', quoting=csv.QUOTE_NONNUMERIC)
        writer.writerow(headers)
    elif fmt == "xls":
        outfile += ".xls"
        book = xlwt.Workbook()
        sheet1 = book.add_sheet("Dados VertNet")
        row1 = sheet1.row(0)
        for i in range(len(headers)):
            row1.write(i, headers[i])
    elif fmt == "xlsx":
        outfile += ".xlsx"
        book = openpyxl.Workbook()
        sheet1 = book.active
        sheet1.title = "Dados VertNet"
        for i in range(1, len(headers) + 1):
            sheet1.cell(row = 1, column = i).value = headers[i - 1]
    elif fmt == "ods":
        outfile += ".ods"
        book = newdoc(doctype="ods", filename=outfile)
        book.sheets.append(Table("Dados VertNet"))
        sheet1 = book.sheets[0]
        sheet1.append_columns(len(headers))
        for i in range(len(headers)):
            sheet1[0, i].set_value(headers[i])
    
    datafile = csv.reader(open(infile, "rb"), delimiter = '\t')
    datafile.next()
    
    recnum = 0
    for row in datafile:
        try:
            DateCollected = row[fields['verbatimeventdate']]
            InstitutionCode = row[fields['institutioncode']]
            CollectionCode = row[fields['collectioncode']]
            CatalogNumber = iif(len(row[fields['catalognumber']]) > 0, row[fields['catalognumber']], row[fields['occurrenceid']])
            CollectorName = unicode_to_ascii(row[fields['recordedby']])
            Sex = row[fields['sex']]
            LifeStage = row[fields['lifestage']]
            ReproductiveCondition = row[fields['reproductivecondition']]
            Behavior = row[fields['behavior']]
            Habitat = row[fields['habitat']]
            CollectorNumber = row[fields['fieldnumber']]
            Notes = row[22] + iif(len(row[fields['fieldnotes']]) > 0, ". " + row[fields['fieldnotes']], '')
            Locality = unicode_to_ascii(row[fields['locality']])
            MinAltitude = to_float(row[fields['minimumelevationinmeters']])
            MaxAltitude = to_float(row[fields['maximumelevationinmeters']])
            MinDepth = to_float(row[fields['minimumdepthinmeters']])
            MaxDepth = to_float(row[fields['maximumdepthinmeters']])
            Latitude = to_float(row[fields['decimallatitude']])
            Longitude = to_float(row[fields['decimallongitude']])
            ScientificName = row[fields['scientificname']]
            Kingdom = row[fields['kingdom']]
            Phylum = row[fields['phylum']]
            Class = row[fields['class']]
            Order = row[fields['order']]
            Family = row[fields['family']]
        except:
            continue
    
        if DateCollected.find('/') != -1:
            DateCollected = DateCollected.replace('/', '-')
    
        if not alfa(InstitutionCode):
            MyCollectionCode = CollectionCode + CatalogNumber
        else:
            MyCollectionCode = InstitutionCode + CatalogNumber
        
        if CollectorName.find("Collector(s):") != -1:
            CollectorName = CollectorName.split(": ")[1]
        
        if MinDepth == 0.0 and MaxDepth == 0.0:
            Elevation = 0.0
        else:
            Elevation = str(MinDepth) + '-' + str(MaxDepth)
        
        if Elevation == 0.0:
            if MinAltitude > 0.0 or MaxAltitude > 0.0:
                Elevation = str(MinAltitude) + '-' + str(MaxAltitude)
                
        if fmt == "csv":
            data = (recnum + 1, MyCollectionCode, Family, ScientificName, \
                    CollectorName, CollectorNumber, DateCollected,\
                    Locality, Latitude, Longitude, Elevation, \
                    Kingdom, Phylum, Class, Order, \
                    Sex, LifeStage, ReproductiveCondition, Behavior, Habitat, Notes)
            writer.writerow(data)
        elif fmt == "xls":
            sheet1.row(recnum + 1).write(0, recnum + 1)
            sheet1.row(recnum + 1).write(1, MyCollectionCode)
            sheet1.row(recnum + 1).write(2, Family)
            sheet1.row(recnum + 1).write(3, ScientificName)
            sheet1.row(recnum + 1).write(4, CollectorName)
            sheet1.row(recnum + 1).write(5, CollectorNumber)
            sheet1.row(recnum + 1).write(6, DateCollected)
            sheet1.row(recnum + 1).write(7, Locality)
            sheet1.row(recnum + 1).write(8, Latitude)
            sheet1.row(recnum + 1).write(9, Longitude)
            sheet1.row(recnum + 1).write(10, Elevation)
            sheet1.row(recnum + 1).write(11, Kingdom)
            sheet1.row(recnum + 1).write(12, Phylum)
            sheet1.row(recnum + 1).write(13, Class)
            sheet1.row(recnum + 1).write(14, Order)
            sheet1.row(recnum + 1).write(15, Sex)
            sheet1.row(recnum + 1).write(16, LifeStage)
            sheet1.row(recnum + 1).write(17, ReproductiveCondition)
            sheet1.row(recnum + 1).write(18, Behavior)
            sheet1.row(recnum + 1).write(19, Habitat)
            sheet1.row(recnum + 1).write(20, Notes)
        elif fmt == "xlsx":
            sheet1.cell(row=recnum + 2, column=1).value = recnum + 1
            sheet1.cell(row=recnum + 2, column=2).value = MyCollectionCode
            sheet1.cell(row=recnum + 2, column=3).value = Family
            sheet1.cell(row=recnum + 2, column=4).value = ScientificName
            sheet1.cell(row=recnum + 2, column=5).value = CollectorName
            sheet1.cell(row=recnum + 2, column=6).value = CollectorNumber
            sheet1.cell(row=recnum + 2, column=7).value = DateCollected
            sheet1.cell(row=recnum + 2, column=8).value = Locality
            sheet1.cell(row=recnum + 2, column=9).value = Latitude
            sheet1.cell(row=recnum + 2, column=10).value = Longitude
            sheet1.cell(row=recnum + 2, column=11).value = Elevation
            sheet1.cell(row=recnum + 2, column=12).value = Kingdom
            sheet1.cell(row=recnum + 2, column=13).value = Phylum
            sheet1.cell(row=recnum + 2, column=14).value = Class
            sheet1.cell(row=recnum + 2, column=15).value = Order
            sheet1.cell(row=recnum + 2, column=16).value = Sex
            sheet1.cell(row=recnum + 2, column=17).value = LifeStage
            sheet1.cell(row=recnum + 2, column=18).value = ReproductiveCondition
            sheet1.cell(row=recnum + 2, column=19).value = Behavior
            sheet1.cell(row=recnum + 2, column=20).value = Habitat
            sheet1.cell(row=recnum + 2, column=21).value = Notes
        elif fmt == "ods":
            sheet1.append_rows(1)
            sheet1[recnum + 1, 0].set_value(recnum + 1)
            sheet1[recnum + 1, 1].set_value(MyCollectionCode)
            sheet1[recnum + 1, 2].set_value(Family)
            sheet1[recnum + 1, 3].set_value(ScientificName)
            sheet1[recnum + 1, 4].set_value(CollectorName)
            sheet1[recnum + 1, 5].set_value(CollectorNumber)
            sheet1[recnum + 1, 6].set_value(DateCollected)
            sheet1[recnum + 1, 7].set_value(Locality)
            sheet1[recnum + 1, 8].set_value(Latitude)
            sheet1[recnum + 1, 9].set_value(Longitude)
            sheet1[recnum + 1, 10].set_value(Elevation)
            sheet1[recnum + 1, 11].set_value(Kingdom)
            sheet1[recnum + 1, 12].set_value(Phylum)
            sheet1[recnum + 1, 13].set_value(Class)
            sheet1[recnum + 1, 14].set_value(Order)
            sheet1[recnum + 1, 15].set_value(Sex)
            sheet1[recnum + 1, 16].set_value(LifeStage)
            sheet1[recnum + 1, 17].set_value(ReproductiveCondition)
            sheet1[recnum + 1, 18].set_value(Behavior)
            sheet1[recnum + 1, 19].set_value(Habitat)
            sheet1[recnum + 1, 20].set_value(Notes)
        recnum += 1

    if fmt == "xls":
        book.save(outfile)
    elif fmt == "xlsx":
        book.save(filename=outfile)
    elif fmt == "ods":
        book.save()
        
    return recnum, outfile