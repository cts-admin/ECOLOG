# -*- coding: utf-8 -*-
#================================================================================#
#  ECOLOG - Sistema Gerenciador de Banco de Dados para Levantamentos Ecológicos  #
#        ECOLOG - Database Management System for Ecological Surveys              #
#      Copyright (c) 1990-2016 Mauro J. Cavalcanti. All rights reserved.         #
#                                                                                #
#   Este programa é software livre; você pode redistribuí-lo e/ou                #
#   modificá-lo sob os termos da Licença Pública Geral GNU, conforme             #
#   publicada pela Free Software Foundation; tanto a versão 2 da                 #
#   Licença como (a seu critério) qualquer versão mais nova.                     #
#                                                                                # 
#   Este programa é distribuído na expectativa de ser útil, mas SEM              #
#   QUALQUER GARANTIA; sem mesmo a garantia implícita de                         #
#   COMERCIALIZAÇÃO ou de ADEQUAÇÃO A QUALQUER PROPÓSITO EM                      #
#   PARTICULAR. Consulte a Licença Pública Geral GNU para obter mais             #
#   detalhes.                                                                    #
#                                                                                #
#   This program is free software: you can redistribute it and/or                #
#   modify it under the terms of the GNU General Public License as published     #
#   by the Free Software Foundation, either version 2 of the License, or         #
#   version 3 of the License, or (at your option) any later version.             #
#                                                                                #
#   This program is distributed in the hope that it will be useful,              #
#   but WITHOUT ANY WARRANTY; without even the implied warranty of               #
#   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See                     #
#   the GNU General Public License for more details.                             #
#                                                                                #
#  Dependências / Dependencies:                                                  #
#    Python 2.6+  (www.python.org)                                               #
#    PyQt 4.8+  (www.riverbankcomputing.com/software/pyqt)                       #
#    NumPy 1.4+ (www.numpy.org)                                                  #
#    xlrd 0.7+ (www.python-excel.org)                                            #
#    openpyxl 2.0+ (openpyxl.readthedocs.org/en/2.0)                             #
#    ezodf 0.2+ (pythonhosted.org/ezodf)                                         #
#    gspread 0.2+ (burnash.github.com/gspread)                                   #
#================================================================================#

import os, sys, csv, sqlite3
import numpy
import gspread
from xlrd import open_workbook
from openpyxl import load_workbook
from ezodf import opendoc
from PyQt4 import QtCore, QtGui

from Useful import (is_number, quote_identifier, substr)

class Sheet(QtGui.QTableWidget):
    def __init__(self, filename=QtCore.QString(), canEdit=False):
        super(QtGui.QTableWidget, self).__init__()
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose)
        self.filename = filename
        self.canEdit = canEdit
        self.data = []
        self.headers = []
        self.types = []
        self.setMinimumSize(500, 300)
        self.title = QtCore.QFileInfo(self.filename).fileName()
        self.setWindowTitle(self.title)
        
    def loadData(self, user=None, pwd=None):    
        
        #--- Excel 97/2000/2003
        if self.filename.endswith(".xls"):
            try:
                wb = open_workbook(self.filename)
                s = wb.sheet_by_index(0)
                for row in range(s.nrows):
                    values = []
                    for col in range(s.ncols):
                        if s.cell(row, col).ctype == 0:
                            val = "" #None
                        else:
                            val = s.cell(row, col).value
                        if row == 0:
                            self.headers.append(val)
                        elif row == 1:
                            if s.cell(row, col).ctype == 1: #or s.cell(row, col).ctype == 3: 
                                self.types.append("text")
                            elif s.cell(row, col).ctype == 2: 
                                self.types.append("numeric")
                            else:
                                self.types.append("text")
                            values.append(val)
                        else:
                            values.append(val)
                    if len(values) > 0: self.data.append(values)
            except (IOError, OSError), e:
                return e
        
        #-- Excel 2007/2010
        elif self.filename.endswith(".xlsx"):
            try:
                wb = load_workbook(filename=self.filename, use_iterators=True)
                s = wb.worksheets[0]
                nrow = 0
                for row in s.iter_rows():
                    values = []
                    for cell in row:
                        val = cell.value
                        if nrow == 0:
                            self.headers.append(val)
                        elif nrow == 1:
                            if cell.data_type == 's': 
                                self.types.append("text")
                            elif cell.data_type == 'n': 
                                self.types.append("numeric")
                            else:
                                self.types.append("text")
                            values.append(val)
                        else:	
                            values.append(val)
                    nrow += 1
                    if len(values) > 0: self.data.append(values)
            except (IOError, OSError), e:
                return e
            
        #-- LibreOffice/OpenOffice
        elif self.filename.endswith(".ods"):
            try:
                wb = opendoc(self.filename)
                s = wb.sheets[0]
                for row in range(s.nrows()-1):
                    values = []
                    for col in range(s.ncols()-1):
                        val = s[(row,col)].value
                        if type(val) == str:
                            val = unicode(val, "utf-8", errors="ignore")
                        else:
                            val = unicode(val)
                        if row == 0:
                            self.headers.append(val)
                        elif row == 1:
                            if s[(row, col)].value_type == "string": #or \
                                #s[(row, col)].value_type == "date" #or \
                                #s[(row, col)].value_type == "time":
                                self.types.append("text")
                            elif s[(row, col)].value_type == "float": 
                                self.types.append("numeric")
                            else:
                                self.types.append("text")
                            values.append(val)
                        else:	
                            values.append(str(val))
                    if len(values) > 0: self.data.append(values)
                del self.data[len(self.data)-1]
            except (IOError, OSError), e:
                return e
            
        #--- Comma-Separated Values
        elif self.filename.endswith(".csv"):
            try:
                reader = csv.reader(open(self.filename))
                rownum = 0
                for row in reader:
                    if rownum == 0:
                        self.headers.append(row)
                    elif rownum == 1:
                        for cell in row:
                            if is_number(cell):
                                self.types.append("numeric")
                            else:
                                self.types.append("text")
                        self.data.append(row)
                    else:
                        self.data.append(row)
                    rownum += 1
                self.headers = self.headers[0]
            except (IOError, OSError), e:
                return e
            
        #--- Tab-Separated Values
        elif self.filename.endswith(".tsv"):
            try:
                reader = csv.reader(open(self.filename), delimiter='\t')
                rownum = 0
                for row in reader:
                    if rownum == 0:
                        self.headers.append(row)
                    elif rownum == 1:
                        for cell in row:
                            if is_number(cell):
                                self.types.append("numeric")
                            else:
                                self.types.append("text")
                        self.data.append(row)
                    else:
                        self.data.append(row)
                    rownum += 1
                self.headers = self.headers[0]
            except (IOError, OSError), e:
                return e
            
        #--- Google Spreadsheet
        elif self.filename.endswith(".docs"):
            try:
                with open(self.filename, 'r') as infile:
                    name = infile.read().strip('\n\r')
                gc = gspread.login(user, pwd)
                sh = gc.open(name)
                ws = sh.get_worksheet(0)
                ws_cells = ws.get_all_values()
                for row in range(ws.row_count):
                    values = []
                    for col in range(ws.col_count):
                        val = ws_cells[row][col]
                        if row == 0:
                            self.headers.append(val)
                        elif row == 1:
                            if is_number(val):
                                self.types.append("numeric")
                            else:
                                self.types.append("text")
                            values.append(val)
                        else:
                            values.append(val)
                    if len(values) > 0: self.data.append(values)
            except (gspread.AuthenticationError, 
                    gspread.SpreadsheetNotFound, 
                    gspread.WorksheetNotFound,
                    gspread.RequestError), e:
                return repr(e)
            
        return None
        
    def setData(self):
        self.clear()
        self.setRowCount(len(self.data))
        self.setColumnCount(len(self.headers))
        for row in range(len(self.data)):
            for col in range(len(self.headers)):
                try:
                    item = str(self.data[row][col])
                    newitem = QtGui.QTableWidgetItem(item)
                    if not self.canEdit:
                        newitem.setFlags(QtCore.Qt.ItemIsSelectable|QtCore.Qt.ItemIsEnabled)
                    self.setItem(row, col, newitem)
                except:
                    pass
        self.setHorizontalHeaderLabels(self.headers)
        for col in range(len(self.headers)):
            self.horizontalHeaderItem(col).setTextAlignment(QtCore.Qt.AlignLeft)
        self.resizeColumnsToContents()
        self.resizeRowsToContents()
        #header = self.horizontalHeader()
        #header.setStretchLastSection(True)
        if not self.canEdit:
            self.setEditTriggers(QtGui.QTableWidget.NoEditTriggers)
            
    def setFilter(self, sqlWhereStr):
        sqlCreateStr = "CREATE TABLE Temp("
        for i in range(len(self.headers)):
            if self.types[i] == "text":
                sqlCreateStr += self.headers[i] + " TEXT"
            elif self.types[i] == "numeric":
                sqlCreateStr += self.headers[i] + " NUMERIC"
            if i < len(self.headers)-1: sqlCreateStr += ", "
        sqlCreateStr = sqlCreateStr + ")"

        sqlInsertStr = "INSERT INTO Temp VALUES("
        for i in range(len(self.headers)):
            sqlInsertStr += "?"
            if i < len(self.headers)-1: sqlInsertStr += ", "
        sqlInsertStr = sqlInsertStr + ")"
        
        sqlQueryStr = "SELECT * FROM Temp "
        if len(sqlWhereStr) > 0: sqlQueryStr += str(sqlWhereStr)

        db_data = tuple(self.data)
        db_connection = sqlite3.connect(":memory:")
        db_cursor = db_connection.cursor()
        db_cursor.execute(sqlCreateStr)
        db_cursor.executemany(sqlInsertStr, db_data)
        db_cursor.execute(sqlQueryStr)
        all_rows = db_cursor.fetchall()
        if len(all_rows) > 0:
            reccount = 0
            self.clear()
            for row in all_rows:
                for col in range(len(self.headers)):
                    item = row[col]
                    newitem = QtGui.QTableWidgetItem(item)
                    newitem.setFlags(QtCore.Qt.ItemIsEnabled)
                    self.setItem(reccount, col, newitem)
                reccount += 1
            self.setHorizontalHeaderLabels(self.headers)
            self.resizeColumnsToContents()
            self.resizeRowsToContents()
        db_cursor.close()
        db_connection.close()
        return(len(all_rows))

    def windowTitle(self):
        return self.title
    
    def gridToArray(self, field1, field2, kind, save_it=False, filename=None, format='i'):
        sqlCreateStr = "CREATE TABLE Temp("
        for i in range(len(self.headers)):
            if self.types[i] == "text":
                sqlCreateStr += quote_identifier(self.headers[i], "replace") + " TEXT"
            elif self.types[i] == "numeric":
                sqlCreateStr += quote_identifier(self.headers[i], "replace") + " NUMERIC"
            if i < len(self.headers) - 1: sqlCreateStr += ", "
        sqlCreateStr += ")"

        sqlInsertStr = "INSERT INTO Temp VALUES("
        for i in range(len(self.headers)):
            sqlInsertStr += "?"
            if i < len(self.headers) - 1: sqlInsertStr += ", "
        sqlInsertStr += ")"
        
        db_data = tuple(self.data)
        db_connection = sqlite3.connect(":memory:")
        db_cursor = db_connection.cursor()
        db_cursor.execute(sqlCreateStr)
        db_cursor.executemany(sqlInsertStr, db_data)

        db_cursor.execute("SELECT DISTINCT " + self.headers[field1] + " FROM Temp")
        rows = db_cursor.fetchall()
        nrows = len(rows)

        if field2 is not None:
            db_cursor.execute("SELECT DISTINCT " + self.headers[field2] + " FROM Temp")
            cols = db_cursor.fetchall()
            ncols = len(cols)
        else:
            sqlSelectStr = "SELECT "
            for i in range(1, len(self.headers)):
                sqlSelectStr += self.headers[i] + ", "
            sqlSelectStr = substr(sqlSelectStr, 0, len(sqlSelectStr) - 2) + " FROM Temp"
            db_cursor.execute(sqlSelectStr)
            cols = db_cursor.fetchall()
            ncols = len(cols[0])
        
        arr = numpy.empty((nrows, ncols), dtype=numpy.float)
        
        if kind in [1, 2]:
            for i in range(nrows):
                row = str(rows[i][0])
                for j in range(ncols):
                    col = cols[j][0]
                    if kind == 1: 
                        db_cursor.execute("SELECT " + self.headers[field2] + " FROM Temp WHERE " + \
                            self.headers[field2] + " = '" + col + "' AND " + \
                            self.headers[field1] + " = '" + row + "'")
                        found = db_cursor.fetchone() != None
                        if found: 
                            bin = 1
                        else:
                            bin = 0
                        arr[i, j] = bin
                    elif kind == 2:	
                        db_cursor.execute("SELECT COUNT(" + self.headers[field2] + ") FROM Temp WHERE " + \
                            self.headers[field2] + " = '" + col + "' AND " + \
                            self.headers[field1] + " = '" + row + "'")
                        freq = db_cursor.fetchone()[0]
                        if freq == None:
                            freq = 0
                        arr[i, j] = freq
        else:
            i = 0
            for col in cols:
                row = list(col)
                j = 0
                for j in range(0,len(row)):
                    try:
                        arr[i, j] = row[j]
                        j += 1
                    except IndexError:
                        continue
                i+= 1
    
        if save_it and filename is not None:
            if format == 'i':
                numpy.savetxt(filename, arr, delimiter=" ", fmt="%5d")
            elif format == 'f':
                numpy.savetxt(filename, arr, delimiter=" ", fmt="%7.2f")
    
        db_cursor.close()
        db_connection.close()
        return arr